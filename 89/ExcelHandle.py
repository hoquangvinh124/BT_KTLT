import openpyxl
from openpyxl import Workbook
from EmployeeClass import Employee

class ExcelHandle:
    @staticmethod
    def read_employees_from_excel(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        employees = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            stt, code, name, age = row
            emp = Employee(str(stt), code, name, str(age))
            employees.append(emp)
        return employees
    @staticmethod
    def write_employees_to_excel(file_path, employees):
        wb = Workbook()
        ws = wb.active
        ws.append(["STT", "Mã", "Tên", "Tuổi"])
        for i, emp in enumerate(employees, start=1):
            ws.append([i, emp.code, emp.name, emp.age])
        wb.save(file_path)

